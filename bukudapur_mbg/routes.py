from __future__ import annotations

from datetime import datetime, timedelta, date
from io import BytesIO
import secrets
import tempfile

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import func

from . import db
from .models import (
    # Access
    AccessCode,
    # Master
    Account,
    Supplier,
    Item,
    # Core
    CashTransaction,
    JournalEntry,
    JournalLine,
    # Sales
    SalesInvoice,
    SalesInvoiceLine,
    ARPayment,
    # Purchase/AP
    Purchase,
    PurchaseItem,
    APayment,
    # Stock usage
    StockUsage,
)
from .pdf_utils import (
    pdf_doc,
    header_block,
    table_2col,
    table_3col,
    table_block,
    fmt_idr,
    footer_canvas,
    section_title,
    subsection_title,
)

# =========================
# Blueprint
# =========================
bp = Blueprint("main", __name__)

# =========================
# Session Keys
# =========================
SESSION_KEY = "access_code"
ADMIN_SESSION_KEY = "admin_logged_in"


# ============================================================
# Helper: Admin
# ============================================================
def _admin_logged_in() -> bool:
    return bool(session.get(ADMIN_SESSION_KEY))


def _require_admin():
    if not _admin_logged_in():
        flash("Silakan login admin dulu.", "error")
        return redirect(url_for("main.admin_login"))
    return None


def _generate_code() -> str:
    return "BDMBG-" + secrets.token_hex(4).upper()


# ============================================================
# Helper: Access / Trial
# ============================================================
def _get_active_access():
    code = session.get(SESSION_KEY)
    if not code:
        return None

    acc = AccessCode.query.filter_by(code=code).first()
    if not acc:
        return None

    changed = acc.mark_expired_if_needed()
    if changed:
        db.session.commit()

    if acc.status == "expired":
        return None

    return acc


def _require_access():
    return _get_active_access()


# ============================================================
# Helper: Tenant scope (per kode akses)
# ============================================================
def _has_col(model_or_alias, col_name: str) -> bool:
    return hasattr(model_or_alias, col_name)


def _scope_filter_for_model(model_or_alias, acc: AccessCode):
    """
    Return filter expression untuk model yang punya kolom access_code_id.
    Kalau tidak ada kolomnya, return None (tidak bisa di-scope).
    """
    if not acc:
        return None
    if _has_col(model_or_alias, "access_code_id"):
        return getattr(model_or_alias, "access_code_id") == acc.id
    return None


def _apply_scope(query, acc: AccessCode, *models_or_aliases):
    """
    Terapkan filter per access_code_id ke query untuk model-model yang punya kolom itu.
    Aman dipakai untuk query join.
    """
    if not acc:
        return query
    for m in models_or_aliases:
        f = _scope_filter_for_model(m, acc)
        if f is not None:
            query = query.filter(f)
    return query


# ============================================================
# Helper: Date parsing + range
# ============================================================
def _parse_date(date_str: str) -> datetime:
    # input HTML date: YYYY-MM-DD
    return datetime.strptime(date_str, "%Y-%m-%d")


def _parse_ymd(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _get_date_range_from_request(default_start_of_month: bool = True):
    """
    Querystring:
      ?from=YYYY-MM-DD&to=YYYY-MM-DD
    Default:
      to = today (UTC date)
      from = awal bulan jika default_start_of_month True, else 30 hari terakhir
    """
    dfrom = _parse_ymd(request.args.get("from"))
    dto = _parse_ymd(request.args.get("to"))

    today = datetime.utcnow().date()
    if dto is None:
        dto = today

    if dfrom is None:
        if default_start_of_month:
            dfrom = dto.replace(day=1)
        else:
            dfrom = dto - timedelta(days=30)

    return dfrom, dto


def _get_date_range_args():
    """
    Ambil query string:
      from / to  (format YYYY-MM-DD)
    Return: (from_dt, to_dt_exclusive, from_str, to_str)
    """
    from_str = (request.args.get("from") or request.args.get("from_date") or "").strip()
    to_str = (request.args.get("to") or request.args.get("to_date") or "").strip()

    from_dt = _parse_date(from_str) if from_str else None
    to_dt_excl = (_parse_date(to_str) + timedelta(days=1)) if to_str else None

    return from_dt, to_dt_excl, from_str, to_str


# ============================================================
# Helper: JournalLine -> JournalEntry FK (robust)
# ============================================================
def _jl_entry_fk():
    """
    Beberapa project pakai JournalLine.entry_id, sebagian journal_entry_id.
    Kita detect otomatis.
    """
    if hasattr(JournalLine, "entry_id"):
        return JournalLine.entry_id
    if hasattr(JournalLine, "journal_entry_id"):
        return JournalLine.journal_entry_id
    raise AttributeError("JournalLine tidak punya entry_id / journal_entry_id")


def _jl_base_query(acc: AccessCode | None, from_dt=None, to_dt_excl=None):
    """
    Base query JournalLine yang JOIN ke JournalEntry (biar bisa filter/order by tanggal).
    + scope per access_code_id kalau kolomnya ada.
    """
    fk = _jl_entry_fk()
    q = JournalLine.query.join(JournalEntry, fk == JournalEntry.id)
    q = _apply_scope(q, acc, JournalEntry, JournalLine)

    if from_dt:
        q = q.filter(JournalEntry.date >= from_dt)
    if to_dt_excl:
        q = q.filter(JournalEntry.date < to_dt_excl)

    return q


# ============================================================
# Helper: account balance (all time / optional date range)
# ============================================================
def _account_balance(
    acc: AccessCode | None,
    code: str,
    from_str: str | None = None,
    to_str: str | None = None,
):
    fk = _jl_entry_fk()
    q = JournalLine.query.join(JournalEntry, fk == JournalEntry.id).filter(JournalLine.account_code == code)
    q = _apply_scope(q, acc, JournalEntry, JournalLine)

    if from_str:
        q = q.filter(JournalEntry.date >= _parse_date(from_str))
    if to_str:
        q = q.filter(JournalEntry.date < (_parse_date(to_str) + timedelta(days=1)))

    debit = q.with_entities(func.coalesce(func.sum(JournalLine.debit), 0.0)).scalar() or 0.0
    credit = q.with_entities(func.coalesce(func.sum(JournalLine.credit), 0.0)).scalar() or 0.0
    return float(debit) - float(credit)


# ============================================================
# Helper: account balance (BY DATE RANGE)
# ============================================================
def _account_balance_range(acc: AccessCode | None, code: str, from_dt=None, to_dt=None):
    """
    Balance debit-credit untuk akun pada rentang tanggal.
    - from_dt/to_dt boleh date (inclusive) atau datetime
    """
    fk = _jl_entry_fk()

    if isinstance(from_dt, date) and not isinstance(from_dt, datetime):
        from_dt = datetime.combine(from_dt, datetime.min.time())

    to_dt_excl = None
    if to_dt is not None:
        if isinstance(to_dt, date) and not isinstance(to_dt, datetime):
            to_dt_excl = datetime.combine(to_dt, datetime.min.time()) + timedelta(days=1)
        else:
            to_dt_excl = to_dt + timedelta(days=1)

    q = (
        db.session.query(
            func.coalesce(func.sum(JournalLine.debit), 0.0).label("debit"),
            func.coalesce(func.sum(JournalLine.credit), 0.0).label("credit"),
        )
        .join(JournalEntry, fk == JournalEntry.id)
        .filter(JournalLine.account_code == code)
    )
    q = _apply_scope(q, acc, JournalEntry, JournalLine)

    if from_dt:
        q = q.filter(JournalEntry.date >= from_dt)
    if to_dt_excl:
        q = q.filter(JournalEntry.date < to_dt_excl)

    row = q.first()
    debit = float(row.debit or 0.0)
    credit = float(row.credit or 0.0)
    return debit - credit


# ============================================================
# Helper: Set scope fields
# ============================================================
def _set_entry_scope(entry: JournalEntry, acc: AccessCode | None):
    if acc and hasattr(entry, "access_code_id"):
        entry.access_code_id = acc.id


def _set_obj_scope(obj, acc: AccessCode | None):
    if acc and hasattr(obj, "access_code_id"):
        obj.access_code_id = acc.id


# ============================================================
# Helper: Jurnal otomatis (scoped)
# ============================================================
def _create_journal_for_cash(acc: AccessCode | None, tx: CashTransaction) -> JournalEntry:
    entry = JournalEntry(date=tx.date, memo=tx.memo, source="cash", source_id=tx.id)
    _set_entry_scope(entry, acc)

    if tx.direction == "in":
        entry.lines.append(
            JournalLine(
                account_code=tx.cash_account_code,
                account_name=tx.cash_account_name,
                debit=tx.amount,
                credit=0,
            )
        )
        entry.lines.append(
            JournalLine(
                account_code=tx.counter_account_code,
                account_name=tx.counter_account_name,
                debit=0,
                credit=tx.amount,
            )
        )
    else:
        entry.lines.append(
            JournalLine(
                account_code=tx.counter_account_code,
                account_name=tx.counter_account_name,
                debit=tx.amount,
                credit=0,
            )
        )
        entry.lines.append(
            JournalLine(
                account_code=tx.cash_account_code,
                account_name=tx.cash_account_name,
                debit=0,
                credit=tx.amount,
            )
        )

    if acc:
        for ln in entry.lines:
            if hasattr(ln, "access_code_id"):
                ln.access_code_id = acc.id

    db.session.add(entry)
    db.session.flush()
    return entry


def _create_journal_for_purchase(acc: AccessCode | None, purchase: Purchase) -> JournalEntry:
    """
    Pembelian hutang:
    Debit Persediaan (10051)
    Kredit Hutang Usaha (20011)
    """
    entry = JournalEntry(date=purchase.date, memo=purchase.memo, source="purchase", source_id=purchase.id)
    _set_entry_scope(entry, acc)
    amount = float(purchase.total_amount or 0)

    inventory_acc = Account.query.filter_by(code="10051").first()
    ap_acc = Account.query.filter_by(code="20011").first()
    if not inventory_acc or not ap_acc:
        raise Exception("Akun Persediaan (10051) atau Hutang Usaha (20011) belum ada.")

    entry.lines.append(
        JournalLine(account_code=inventory_acc.code, account_name=inventory_acc.name, debit=amount, credit=0)
    )
    entry.lines.append(
        JournalLine(account_code=ap_acc.code, account_name=ap_acc.name, debit=0, credit=amount)
    )

    if acc:
        for ln in entry.lines:
            if hasattr(ln, "access_code_id"):
                ln.access_code_id = acc.id

    db.session.add(entry)
    db.session.flush()
    return entry


def _create_journal_for_ap_payment(acc: AccessCode | None, payment: APayment) -> JournalEntry:
    """
    Bayar hutang:
    Debit Hutang Usaha (20011)
    Kredit Kas/Bank (dipilih)
    """
    entry = JournalEntry(date=payment.date, memo=payment.memo, source="ap_payment", source_id=payment.id)
    _set_entry_scope(entry, acc)

    ap_acc = Account.query.filter_by(code="20011").first()
    cash_acc = Account.query.filter_by(code=payment.cash_account_code).first()
    if not ap_acc or not cash_acc:
        raise Exception("Akun Hutang Usaha atau Kas/Bank tidak ditemukan.")

    entry.lines.append(
        JournalLine(account_code=ap_acc.code, account_name=ap_acc.name, debit=float(payment.amount or 0), credit=0)
    )
    entry.lines.append(
        JournalLine(account_code=cash_acc.code, account_name=cash_acc.name, debit=0, credit=float(payment.amount or 0))
    )

    if acc:
        for ln in entry.lines:
            if hasattr(ln, "access_code_id"):
                ln.access_code_id = acc.id

    db.session.add(entry)
    db.session.flush()
    return entry


def _create_journal_for_stock_usage(acc: AccessCode | None, u: StockUsage) -> JournalEntry:
    """
    Pemakaian stok:
    Debit HPP (dipilih)
    Kredit Persediaan (10051)
    """
    inv_acc = Account.query.filter_by(code="10051").first()
    hpp_acc = Account.query.filter_by(code=u.hpp_account_code).first()
    if not inv_acc or not hpp_acc:
        raise Exception("Akun Persediaan (10051) atau akun HPP tidak ditemukan.")

    entry = JournalEntry(date=u.date, memo=u.memo, source="stock_usage", source_id=u.id)
    _set_entry_scope(entry, acc)

    entry.lines.append(
        JournalLine(account_code=hpp_acc.code, account_name=hpp_acc.name, debit=float(u.total_cost or 0), credit=0)
    )
    entry.lines.append(
        JournalLine(account_code=inv_acc.code, account_name=inv_acc.name, debit=0, credit=float(u.total_cost or 0))
    )

    if acc:
        for ln in entry.lines:
            if hasattr(ln, "access_code_id"):
                ln.access_code_id = acc.id

    db.session.add(entry)
    db.session.flush()
    return entry


def _next_invoice_no(prefix="INV"):
    today = datetime.utcnow().strftime("%Y%m%d")
    base = f"{prefix}-{today}-"
    last = (
        SalesInvoice.query.filter(SalesInvoice.invoice_no.like(base + "%"))
        .order_by(SalesInvoice.id.desc())
        .first()
    )
    if not last:
        return base + "001"
    try:
        seq = int(last.invoice_no.split("-")[-1]) + 1
    except Exception:
        seq = 1
    return base + f"{seq:03d}"


def _create_journal_for_invoice(acc: AccessCode | None, inv: SalesInvoice) -> JournalEntry:
    entry = JournalEntry(
        date=inv.date,
        memo=f"Invoice {inv.invoice_no} - {inv.customer_name}",
        source="sales_invoice",
        source_id=inv.id,
    )
    _set_entry_scope(entry, acc)

    entry.lines.append(
        JournalLine(
            account_code=inv.ar_account_code,
            account_name=inv.ar_account_name,
            debit=float(inv.total_amount or 0),
            credit=0,
        )
    )
    entry.lines.append(
        JournalLine(
            account_code=inv.revenue_account_code,
            account_name=inv.revenue_account_name,
            debit=0,
            credit=float(inv.total_amount or 0),
        )
    )

    if acc:
        for ln in entry.lines:
            if hasattr(ln, "access_code_id"):
                ln.access_code_id = acc.id

    db.session.add(entry)
    db.session.flush()
    return entry


def _create_journal_for_ar_payment(acc: AccessCode | None, p: ARPayment, inv: SalesInvoice) -> JournalEntry:
    entry = JournalEntry(
        date=p.date,
        memo=f"Pelunasan {inv.invoice_no} - {inv.customer_name}",
        source="ar_payment",
        source_id=p.id,
    )
    _set_entry_scope(entry, acc)

    entry.lines.append(
        JournalLine(
            account_code=p.cash_account_code,
            account_name=p.cash_account_name,
            debit=float(p.amount or 0),
            credit=0,
        )
    )
    entry.lines.append(
        JournalLine(
            account_code=inv.ar_account_code,
            account_name=inv.ar_account_name,
            debit=0,
            credit=float(p.amount or 0),
        )
    )

    if acc:
        for ln in entry.lines:
            if hasattr(ln, "access_code_id"):
                ln.access_code_id = acc.id

    db.session.add(entry)
    db.session.flush()
    return entry


def _arpay_memo(customer: str | None, note: str | None) -> str:
    cust = (customer or "").strip()
    note = (note or "").strip()
    if cust and note:
        return f"[AR] {cust} - {note}"
    if cust:
        return f"[AR] {cust}"
    if note:
        return f"[AR] {note}"
    return "[AR]"


# ============================================================
# Admin Routes
# ============================================================
@bp.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        pin = (request.form.get("pin") or "").strip()
        if pin == current_app.config.get("ADMIN_PIN"):
            session[ADMIN_SESSION_KEY] = True
            flash("Login admin berhasil.", "success")
            return redirect(url_for("main.admin_codes"))
        flash("PIN salah.", "error")
        return redirect(url_for("main.admin_login"))

    return render_template("admin_login.html")


@bp.post("/admin/logout")
def admin_logout():
    session.pop(ADMIN_SESSION_KEY, None)
    flash("Logout admin.", "success")
    return redirect(url_for("main.admin_login"))


@bp.get("/admin/codes")
def admin_codes():
    guard = _require_admin()
    if guard:
        return guard

    codes = AccessCode.query.order_by(AccessCode.id.desc()).limit(200).all()
    return render_template("admin_codes.html", codes=codes)


@bp.post("/admin/codes/create")
def admin_create_code():
    guard = _require_admin()
    if guard:
        return guard

    dapur_name = (request.form.get("dapur_name") or "").strip()
    status = (request.form.get("status") or "active").strip()
    days_str = (request.form.get("days") or "30").strip()

    try:
        days = int(days_str)
        if days <= 0:
            raise ValueError()
    except ValueError:
        flash("Durasi hari harus angka > 0.", "error")
        return redirect(url_for("main.admin_codes"))

    code = _generate_code()
    start_at = datetime.utcnow()
    expires_at = start_at + timedelta(days=days)

    acc = AccessCode(
        code=code,
        dapur_name=dapur_name or None,
        status=status,
        start_at=start_at,
        expires_at=expires_at,
    )
    db.session.add(acc)
    db.session.commit()

    flash(f"Kode dibuat: {code} (exp: {expires_at})", "success")
    return redirect(url_for("main.admin_codes"))


@bp.post("/admin/codes/extend")
def admin_extend_code():
    guard = _require_admin()
    if guard:
        return guard

    code = (request.form.get("code") or "").strip().upper()
    days_str = (request.form.get("days") or "").strip()

    try:
        days = int(days_str)
        if days <= 0:
            raise ValueError()
    except ValueError:
        flash("Hari perpanjangan harus angka > 0.", "error")
        return redirect(url_for("main.admin_codes"))

    acc = AccessCode.query.filter_by(code=code).first()
    if not acc:
        flash("Kode tidak ditemukan.", "error")
        return redirect(url_for("main.admin_codes"))

    now = datetime.utcnow()
    base = acc.expires_at if (acc.expires_at and acc.expires_at > now) else now
    acc.expires_at = base + timedelta(days=days)
    acc.status = "active"
    if not acc.start_at:
        acc.start_at = now

    db.session.commit()
    flash(f"Kode {acc.code} diperpanjang +{days} hari. Exp: {acc.expires_at}", "success")
    return redirect(url_for("main.admin_codes"))


@bp.post("/admin/codes/expire")
def admin_expire_code():
    guard = _require_admin()
    if guard:
        return guard

    code = (request.form.get("code") or "").strip().upper()
    acc = AccessCode.query.filter_by(code=code).first()
    if not acc:
        flash("Kode tidak ditemukan.", "error")
        return redirect(url_for("main.admin_codes"))

    acc.status = "expired"
    acc.expires_at = datetime.utcnow()
    db.session.commit()

    flash(f"Kode {acc.code} di-expire.", "success")
    return redirect(url_for("main.admin_codes"))


# ============================================================
# Home/Auth
# ============================================================
@bp.get("/")
def home():
    acc = _get_active_access()
    if not acc:
        return redirect(url_for("main.enter_code"))
    return redirect(url_for("main.dashboard"))


@bp.route("/enter", methods=["GET", "POST"])
def enter_code():
    if request.method == "POST":
        code = (request.form.get("code") or "").strip().upper()
        if not code:
            flash("Masukkan kode akses.", "error")
            return redirect(url_for("main.enter_code"))

        acc = AccessCode.query.filter_by(code=code).first()
        if not acc:
            flash("Kode tidak ditemukan.", "error")
            return redirect(url_for("main.enter_code"))

        if acc.mark_expired_if_needed():
            db.session.commit()

        session[SESSION_KEY] = acc.code

        if acc.status == "expired":
            return redirect(url_for("main.expired"))

        flash("Akses berhasil.", "success")
        return redirect(url_for("main.dashboard"))

    return render_template("enter_code.html")


@bp.post("/trial")
def create_trial():
    dapur_name = (request.form.get("dapur_name") or "").strip()

    code = _generate_code()
    start_at = datetime.utcnow()
    expires_at = start_at + timedelta(days=3)

    acc = AccessCode(
        code=code,
        dapur_name=dapur_name or None,
        status="trial",
        start_at=start_at,
        expires_at=expires_at,
    )
    db.session.add(acc)
    db.session.commit()

    session[SESSION_KEY] = code
    flash(f"Trial dibuat. Kode akses kamu: {code}", "success")
    return redirect(url_for("main.dashboard"))


@bp.get("/dashboard")
def dashboard():
    acc = _get_active_access()
    if not acc:
        if session.get(SESSION_KEY):
            return redirect(url_for("main.expired"))
        return redirect(url_for("main.enter_code"))

    remaining = acc.expires_at - datetime.utcnow()
    remaining_hours = max(0, int(remaining.total_seconds() // 3600))

    # ALL-TIME range
    dfrom = datetime(2000, 1, 1)
    now = datetime.utcnow()
    dto = datetime(now.year, now.month, now.day, 23, 59, 59)

    def bal(code: str) -> float:
        return float(_account_balance_range(acc, code, dfrom, dto))

    def sum_by_type(t: str) -> float:
        accs = Account.query.filter(Account.type == t).all()
        total = 0.0
        for a in accs:
            b = bal(a.code)
            if t in ("Pendapatan", "Pendapatan Lain"):
                total += -b
            else:
                total += b
        return float(total)

    rev_main = sum_by_type("Pendapatan")
    hpp = sum_by_type("HPP")
    op_exp = sum_by_type("Beban")
    rev_other = sum_by_type("Pendapatan Lain")
    exp_other = sum_by_type("Beban Lain")

    gross_profit = rev_main - hpp
    operating_profit = gross_profit - op_exp
    net_profit = operating_profit + rev_other - exp_other

    # Top Beban Operasional
    exp_accounts = Account.query.filter(Account.type == "Beban").all()
    tmp = []
    for a in exp_accounts:
        amt = bal(a.code)
        if amt and amt > 0:
            tmp.append((a.name, float(amt)))
    tmp.sort(key=lambda x: x[1], reverse=True)
    tmp = tmp[:5]
    top_exp_labels = [x[0] for x in tmp]
    top_exp_values = [x[1] for x in tmp]

    # Kas & Bank
    cash_accounts = Account.query.filter(Account.type == "Kas & Bank").order_by(Account.code.asc()).all()
    cash_labels = []
    cash_values = []
    cash_total = 0.0
    for a in cash_accounts:
        b = bal(a.code)
        cash_labels.append(f"{a.code} {a.name}")
        cash_values.append(float(b))
        cash_total += float(b)

    # Pie Chart
    chart_pl_labels = [
        "Pendapatan Usaha",
        "HPP",
        "Beban Operasional",
        "Pend. Luar Usaha",
        "Beban Luar Usaha",
        "Laba Bersih",
    ]
    chart_pl_values = [
        abs(float(rev_main)),
        abs(float(hpp)),
        abs(float(op_exp)),
        abs(float(rev_other)),
        abs(float(exp_other)),
        abs(float(net_profit)),
    ]

    return render_template(
        "dashboard.html",
        access=acc,
        remaining_hours=remaining_hours,
        cash_total=cash_total,
        rev_main=rev_main,
        hpp=hpp,
        op_exp=op_exp,
        net_profit=net_profit,
        chart_pl={"labels": chart_pl_labels, "values": chart_pl_values},
        chart_top_exp={"labels": top_exp_labels, "values": [abs(x) for x in top_exp_values]},
        chart_cash={"labels": cash_labels, "values": cash_values},
    )


@bp.get("/expired")
def expired():
    code = session.get(SESSION_KEY)
    acc = AccessCode.query.filter_by(code=code).first() if code else None
    return render_template("expired.html", access=acc)


@bp.post("/logout")
def logout():
    session.pop(SESSION_KEY, None)
    flash("Keluar.", "success")
    return redirect(url_for("main.enter_code"))


# ============================================================
# PART 2/4 — MASTER DATA (scoped per access_code_id)
# ============================================================

# ============================================================
# Master Data Home
# ============================================================
@bp.get("/master")
def master_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))
    return render_template("master_home.html")


# ============================================================
# Master: Accounts (COA)
# ============================================================
@bp.route("/master/accounts", methods=["GET", "POST"])
def master_accounts():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    if request.method == "POST":
        code = (request.form.get("code") or "").strip()
        name = (request.form.get("name") or "").strip()
        atype = (request.form.get("type") or "").strip()

        if not code or not name or not atype:
            flash("Kode, Nama, dan Tipe akun wajib diisi.", "error")
            return redirect(url_for("main.master_accounts"))

        # ✅ unique per dapur (access_code_id)
        exists = Account.query.filter_by(access_code_id=acc.id, code=code).first()
        if exists:
            flash("Kode akun sudah ada di dapur ini.", "error")
            return redirect(url_for("main.master_accounts"))

        obj = Account(access_code_id=acc.id, code=code, name=name, type=atype)
        db.session.add(obj)
        db.session.commit()

        flash("Akun berhasil ditambahkan.", "success")
        return redirect(url_for("main.master_accounts"))

    accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .order_by(Account.code.asc())
        .all()
    )
    return render_template("master_accounts.html", accounts=accounts)


@bp.post("/master/accounts/seed")
def seed_accounts():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    standard_accounts = [
        ("10011", "Kas", "Kas & Bank"),
        ("10021", "Bank", "Kas & Bank"),
        ("10031", "Piutang Usaha", "Akun Piutang"),
        ("10041", "Piutang Karyawan", "Aktiva Lancar Lain"),
        ("10042", "PPN Masukan", "Aktiva Lancar Lain"),
        ("10051", "Persediaan", "Persediaan"),
        ("10061", "Peralatan", "Aktiva Tetap"),
        ("10071", "Akum. Penyusutan Peralatan", "Akum. Peny."),
        ("20011", "Hutang Usaha", "Akun Hutang"),
        ("20021", "Hutang Lain", "Hutang Lancar Lain"),
        ("20022", "PPN Keluaran", "Hutang Lancar Lain"),
        ("20031", "Hutang Bank", "Hutang Jk. Panjang"),
        ("30011", "Modal", "Ekuitas"),
        ("30021", "Prive/Deviden", "Ekuitas"),
        ("30031", "Laba Ditahan", "Ekuitas"),
        ("40011", "Penjualan", "Pendapatan"),
        ("50011", "Beban Pokok Dapur", "HPP"),
        ("60011", "Biaya Gaji & Upah", "Beban"),
        ("60012", "Biaya Listrik", "Beban"),
        ("60013", "Biaya Promosi", "Beban"),
        ("60014", "Biaya Komisi", "Beban"),
        ("60015", "Biaya Perlengkapan Dapur", "Beban"),
        ("60016", "Biaya ATK", "Beban"),
        ("60017", "Biaya Pengiriman", "Beban"),
        ("60018", "Biaya Transportasi", "Beban"),
        ("60019", "Biaya Legalitas & Perizinan", "Beban"),
        ("60020", "Biaya PAM", "Beban"),
        ("60021", "Biaya Kebersihan Keamanan", "Beban"),
        ("60022", "Biaya Pajak", "Beban"),
        ("60099", "Biaya Lain-lain", "Beban"),
        ("70011", "Pendapatan Bunga Bank", "Pendapatan Lain"),
        ("80011", "Biaya Adm Bank", "Beban Lain"),
    ]

    inserted = 0
    skipped = 0

    for code, name, atype in standard_accounts:
        exists = Account.query.filter_by(access_code_id=acc.id, code=code).first()
        if exists:
            skipped += 1
            continue

        db.session.add(
            Account(access_code_id=acc.id, code=code, name=name, type=atype)
        )
        inserted += 1

    db.session.commit()
    flash(f"Import akun standar selesai. Ditambah: {inserted}, dilewati: {skipped}.", "success")
    return redirect(url_for("main.master_accounts"))


# ============================================================
# Master: Suppliers
# ============================================================
@bp.route("/master/suppliers", methods=["GET", "POST"])
def master_suppliers():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        phone = (request.form.get("phone") or "").strip()
        address = (request.form.get("address") or "").strip()

        if not name:
            flash("Nama supplier wajib diisi.", "error")
            return redirect(url_for("main.master_suppliers"))

        obj = Supplier(
            access_code_id=acc.id,
            name=name,
            phone=phone or None,
            address=address or None,
        )
        db.session.add(obj)
        db.session.commit()

        flash("Supplier berhasil ditambahkan.", "success")
        return redirect(url_for("main.master_suppliers"))

    suppliers = (
        Supplier.query.filter_by(access_code_id=acc.id)
        .order_by(Supplier.name.asc())
        .all()
    )
    return render_template("master_suppliers.html", suppliers=suppliers)


# ============================================================
# Master: Items
# ============================================================
@bp.route("/master/items", methods=["GET", "POST"])
def master_items():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        category = (request.form.get("category") or "").strip()
        unit = (request.form.get("unit") or "").strip()
        min_stock = (request.form.get("min_stock") or "0").strip()

        if not name or not unit:
            flash("Nama bahan dan satuan wajib diisi.", "error")
            return redirect(url_for("main.master_items"))

        try:
            min_stock_val = float(min_stock)
            if min_stock_val < 0:
                raise ValueError()
        except ValueError:
            flash("Minimal stok harus angka >= 0.", "error")
            return redirect(url_for("main.master_items"))

        obj = Item(
            access_code_id=acc.id,
            name=name,
            category=category or None,
            unit=unit,
            min_stock=min_stock_val,
        )
        db.session.add(obj)
        db.session.commit()

        flash("Bahan berhasil ditambahkan.", "success")
        return redirect(url_for("main.master_items"))

    items = (
        Item.query.filter_by(access_code_id=acc.id)
        .order_by(Item.name.asc())
        .all()
    )
    return render_template("master_items.html", items=items)

# ============================================================
# PART 3/4 — TRANSAKSI (scoped per access_code_id)
#   - Cash
#   - Journals list/detail
#   - Purchase + helpers edit/delete
#   - AP Payment + edit/delete
#   - Sales (simple tag [SALE]) + edit/delete
#   - AR Payments (invoice) + (opsional) AR Settlement via CashTransaction memo [AR]
#   - Expenses + edit/delete
#   - Stock Usage + edit/delete
# ============================================================

# ============================================================
# Kas
# ============================================================
@bp.route("/cash", methods=["GET", "POST"])
def cash_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        direction = (request.form.get("direction") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        counter_code = (request.form.get("counter_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or direction not in ("in", "out") or not cash_code or not counter_code or not amount_str:
            flash("Tanggal, tipe, akun kas/bank, akun lawan, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.cash_home"))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.cash_home"))

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        counter_acc = Account.query.filter_by(access_code_id=acc.id, code=counter_code).first()
        if not cash_acc or not counter_acc:
            flash("Akun tidak valid. Pastikan sudah ada di COA.", "error")
            return redirect(url_for("main.cash_home"))

        tx = CashTransaction(
            access_code_id=acc.id,
            date=_parse_date(date_str),
            direction=direction,
            cash_account_code=cash_acc.code,
            cash_account_name=cash_acc.name,
            counter_account_code=counter_acc.code,
            counter_account_name=counter_acc.name,
            amount=amount,
            memo=memo or None,
        )
        db.session.add(tx)
        db.session.flush()

        entry = _create_journal_for_cash(tx)
        tx.journal_entry_id = entry.id

        db.session.commit()
        flash("Transaksi kas tersimpan & jurnal otomatis dibuat.", "success")
        return redirect(url_for("main.cash_home"))

    txs = (
        CashTransaction.query.filter_by(access_code_id=acc.id)
        .order_by(CashTransaction.date.desc(), CashTransaction.id.desc())
        .limit(50)
        .all()
    )
    return render_template("cash_home.html", accounts=accounts, txs=txs)


@bp.route("/cash/<int:tx_id>/edit", methods=["GET", "POST"])
def cash_edit(tx_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    tx = CashTransaction.query.filter_by(id=tx_id, access_code_id=acc.id).first_or_404()
    accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        direction = (request.form.get("direction") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        counter_code = (request.form.get("counter_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or direction not in ("in", "out") or not cash_code or not counter_code or not amount_str:
            flash("Tanggal, tipe, akun kas/bank, akun lawan, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.cash_edit", tx_id=tx_id))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.cash_edit", tx_id=tx_id))

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        counter_acc = Account.query.filter_by(access_code_id=acc.id, code=counter_code).first()
        if not cash_acc or not counter_acc:
            flash("Akun tidak valid.", "error")
            return redirect(url_for("main.cash_edit", tx_id=tx_id))

        # 1) hapus JournalEntry lama (kalau ada)
        if getattr(tx, "journal_entry_id", None):
            old_entry = JournalEntry.query.filter_by(id=tx.journal_entry_id, access_code_id=acc.id).first()
            if old_entry:
                db.session.delete(old_entry)
                db.session.flush()
            tx.journal_entry_id = None

        # 2) update transaksi
        tx.date = _parse_date(date_str)
        tx.direction = direction
        tx.cash_account_code = cash_acc.code
        tx.cash_account_name = cash_acc.name
        tx.counter_account_code = counter_acc.code
        tx.counter_account_name = counter_acc.name
        tx.amount = amount
        tx.memo = memo or None

        db.session.flush()

        # 3) buat ulang jurnal
        entry = _create_journal_for_cash(tx)
        tx.journal_entry_id = entry.id

        db.session.commit()
        flash("Transaksi kas berhasil diupdate.", "success")
        return redirect(url_for("main.cash_home"))

    return render_template("cash_edit.html", tx=tx, accounts=accounts)


@bp.post("/cash/<int:tx_id>/delete")
def cash_delete(tx_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    tx = CashTransaction.query.filter_by(id=tx_id, access_code_id=acc.id).first_or_404()

    # hapus journal entry terkait
    if getattr(tx, "journal_entry_id", None):
        entry = JournalEntry.query.filter_by(id=tx.journal_entry_id, access_code_id=acc.id).first()
        if entry:
            db.session.delete(entry)

    db.session.delete(tx)
    db.session.commit()
    flash("Transaksi kas berhasil dihapus.", "success")
    return redirect(url_for("main.cash_home"))


# ============================================================
# Jurnal (dengan filter tanggal) — scoped
# ============================================================
@bp.get("/journals")
def journals_list():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    dfrom, dto = _get_date_range_from_request()

    entries = (
        JournalEntry.query.filter_by(access_code_id=acc.id)
        .filter(JournalEntry.date >= datetime.combine(dfrom, datetime.min.time()))
        .filter(JournalEntry.date <= datetime.combine(dto, datetime.max.time()))
        .order_by(JournalEntry.date.desc(), JournalEntry.id.desc())
        .limit(200)
        .all()
    )
    return render_template(
        "journals_list.html",
        entries=entries,
        dfrom=dfrom.strftime("%Y-%m-%d"),
        dto=dto.strftime("%Y-%m-%d"),
    )


@bp.get("/journals/<int:entry_id>")
def journals_detail(entry_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    entry = JournalEntry.query.filter_by(id=entry_id, access_code_id=acc.id).first_or_404()
    return render_template("journals_detail.html", entry=entry)


# ============================================================
# Purchase (hutang) — scoped
# ============================================================
@bp.route("/purchase", methods=["GET", "POST"])
def purchase_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    suppliers = (
        Supplier.query.filter_by(access_code_id=acc.id)
        .order_by(Supplier.name.asc())
        .all()
    )
    items = (
        Item.query.filter_by(access_code_id=acc.id)
        .order_by(Item.name.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        supplier_id = (request.form.get("supplier_id") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        item_id = (request.form.get("item_id") or "").strip()
        qty_str = (request.form.get("qty") or "").strip()
        price_str = (request.form.get("price") or "").strip()

        if not date_str or not item_id or not qty_str or not price_str:
            flash("Tanggal, bahan, qty, dan harga wajib diisi.", "error")
            return redirect(url_for("main.purchase_home"))

        try:
            qty = float(qty_str)
            price = float(price_str)
            if qty <= 0 or price <= 0:
                raise ValueError()
        except ValueError:
            flash("Qty dan harga harus angka > 0.", "error")
            return redirect(url_for("main.purchase_home"))

        item = Item.query.filter_by(id=int(item_id), access_code_id=acc.id).first()
        if not item:
            flash("Bahan tidak valid.", "error")
            return redirect(url_for("main.purchase_home"))

        subtotal = qty * price

        purchase = Purchase(
            access_code_id=acc.id,
            date=_parse_date(date_str),
            total_amount=subtotal,
            memo=memo or None,
        )

        if supplier_id:
            supplier = Supplier.query.filter_by(id=int(supplier_id), access_code_id=acc.id).first()
            if supplier:
                purchase.supplier_id = supplier.id
                purchase.supplier_name = supplier.name

        db.session.add(purchase)
        db.session.flush()

        pitem = PurchaseItem(
            access_code_id=acc.id,
            purchase_id=purchase.id,
            item_id=item.id,
            item_name=item.name,
            qty=qty,
            price=price,
            subtotal=subtotal,
        )
        db.session.add(pitem)

        # update stok & avg cost (moving average)
        total_cost_existing = float(item.stock_qty or 0) * float(item.avg_cost or 0)
        total_cost_new = qty * price
        new_qty = float(item.stock_qty or 0) + qty

        item.avg_cost = (total_cost_existing + total_cost_new) / new_qty
        item.stock_qty = new_qty

        entry = _create_journal_for_purchase(purchase)
        purchase.journal_entry_id = entry.id

        db.session.commit()
        flash("Pembelian tersimpan, stok bertambah, hutang tercatat.", "success")
        return redirect(url_for("main.purchase_home"))

    purchases = (
        Purchase.query.filter_by(access_code_id=acc.id)
        .order_by(Purchase.date.desc(), Purchase.id.desc())
        .limit(20)
        .all()
    )
    return render_template("purchase_home.html", suppliers=suppliers, items=items, purchases=purchases)


# ============================================================
# PURCHASE: Helpers reverse/apply stok + rebuild jurnal (scoped)
# ============================================================
def _reverse_purchase_stock(acc: AccessCode, pitem: PurchaseItem):
    item = Item.query.filter_by(id=pitem.item_id, access_code_id=acc.id).first()
    if not item:
        return

    old_qty = float(pitem.qty or 0)
    old_price = float(pitem.price or 0)
    if old_qty <= 0:
        return

    cur_qty = float(item.stock_qty or 0)
    cur_avg = float(item.avg_cost or 0)

    total_cost_cur = cur_qty * cur_avg
    total_cost_old = old_qty * old_price

    new_qty = cur_qty - old_qty
    if new_qty <= 0:
        item.stock_qty = 0.0
        item.avg_cost = 0.0
        return

    new_total_cost = total_cost_cur - total_cost_old
    if new_total_cost < 0:
        new_total_cost = 0.0

    item.stock_qty = new_qty
    item.avg_cost = new_total_cost / new_qty if new_qty else 0.0


def _apply_purchase_stock(item: Item, qty: float, price: float):
    qty = float(qty or 0)
    price = float(price or 0)
    if qty <= 0:
        return

    cur_qty = float(item.stock_qty or 0)
    cur_avg = float(item.avg_cost or 0)

    total_cost_existing = cur_qty * cur_avg
    total_cost_new = qty * price
    new_qty = cur_qty + qty

    item.stock_qty = new_qty
    item.avg_cost = (total_cost_existing + total_cost_new) / new_qty if new_qty else 0.0


def _delete_journal_entry_scoped(acc: AccessCode, entry_id: int | None):
    if not entry_id:
        return
    # hapus lines dulu (scoped)
    JournalLine.query.filter_by(access_code_id=acc.id, entry_id=entry_id).delete()
    JournalEntry.query.filter_by(access_code_id=acc.id, id=entry_id).delete()


def _rebuild_journal_for_purchase(acc: AccessCode, purchase: Purchase):
    _delete_journal_entry_scoped(acc, getattr(purchase, "journal_entry_id", None))
    db.session.flush()
    entry = _create_journal_for_purchase(purchase)
    purchase.journal_entry_id = entry.id


# ============================================================
# PURCHASE: Edit / Delete (scoped)
# ============================================================
@bp.route("/purchase/<int:purchase_id>/edit", methods=["GET", "POST"])
def purchase_edit(purchase_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    purchase = Purchase.query.filter_by(id=purchase_id, access_code_id=acc.id).first_or_404()

    pitem = PurchaseItem.query.filter_by(
        access_code_id=acc.id,
        purchase_id=purchase.id,
    ).first()
    if not pitem:
        flash("Item pembelian tidak ditemukan.", "error")
        return redirect(url_for("main.purchase_home"))

    suppliers = Supplier.query.filter_by(access_code_id=acc.id).order_by(Supplier.name.asc()).all()
    items = Item.query.filter_by(access_code_id=acc.id).order_by(Item.name.asc()).all()

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        supplier_id = (request.form.get("supplier_id") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        item_id = (request.form.get("item_id") or "").strip()
        qty_str = (request.form.get("qty") or "").strip()
        price_str = (request.form.get("price") or "").strip()

        if not date_str or not item_id or not qty_str or not price_str:
            flash("Tanggal, bahan, qty, dan harga wajib diisi.", "error")
            return redirect(url_for("main.purchase_edit", purchase_id=purchase.id))

        try:
            qty = float(qty_str)
            price = float(price_str)
            if qty <= 0 or price <= 0:
                raise ValueError()
        except ValueError:
            flash("Qty dan harga harus angka > 0.", "error")
            return redirect(url_for("main.purchase_edit", purchase_id=purchase.id))

        new_item = Item.query.filter_by(id=int(item_id), access_code_id=acc.id).first()
        if not new_item:
            flash("Bahan tidak valid.", "error")
            return redirect(url_for("main.purchase_edit", purchase_id=purchase.id))

        # STEP 1: reverse stok dari pembelian lama
        _reverse_purchase_stock(acc, pitem)

        # STEP 2: update purchase + pitem
        purchase.date = _parse_date(date_str)
        purchase.memo = memo or None

        if supplier_id:
            sup = Supplier.query.filter_by(id=int(supplier_id), access_code_id=acc.id).first()
            if sup:
                purchase.supplier_id = sup.id
                purchase.supplier_name = sup.name
            else:
                purchase.supplier_id = None
                purchase.supplier_name = None
        else:
            purchase.supplier_id = None
            purchase.supplier_name = None

        pitem.item_id = new_item.id
        pitem.item_name = new_item.name
        pitem.qty = qty
        pitem.price = price
        pitem.subtotal = qty * price
        purchase.total_amount = pitem.subtotal

        # STEP 3: apply stok baru
        _apply_purchase_stock(new_item, qty, price)

        # STEP 4: rebuild jurnal pembelian
        _rebuild_journal_for_purchase(acc, purchase)

        db.session.commit()
        flash("Pembelian berhasil diupdate. Stok & jurnal sudah disesuaikan.", "success")
        return redirect(url_for("main.purchase_home"))

    return render_template(
        "purchase_edit.html",
        purchase=purchase,
        pitem=pitem,
        suppliers=suppliers,
        items=items,
    )


@bp.post("/purchase/<int:purchase_id>/delete")
def purchase_delete(purchase_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    purchase = Purchase.query.filter_by(id=purchase_id, access_code_id=acc.id).first_or_404()
    pitem = PurchaseItem.query.filter_by(access_code_id=acc.id, purchase_id=purchase.id).first()

    if pitem:
        _reverse_purchase_stock(acc, pitem)

    _delete_journal_entry_scoped(acc, getattr(purchase, "journal_entry_id", None))

    if pitem:
        db.session.delete(pitem)

    db.session.delete(purchase)
    db.session.commit()

    flash("Pembelian dihapus. Stok & jurnal sudah dikembalikan.", "success")
    return redirect(url_for("main.purchase_home"))


# ============================================================
# AP Payment (scoped)
# ============================================================
@bp.route("/ap-payment", methods=["GET", "POST"])
def ap_payment_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    purchases = (
        Purchase.query.filter_by(access_code_id=acc.id)
        .order_by(Purchase.date.desc(), Purchase.id.desc())
        .all()
    )
    cash_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type == "Kas & Bank")
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        purchase_id = (request.form.get("purchase_id") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not cash_code or not amount_str:
            flash("Tanggal, akun kas, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.ap_payment_home"))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.ap_payment_home"))

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        if not cash_acc:
            flash("Akun kas/bank tidak valid.", "error")
            return redirect(url_for("main.ap_payment_home"))

        payment = APayment(
            access_code_id=acc.id,
            date=_parse_date(date_str),
            amount=amount,
            cash_account_code=cash_acc.code,
            cash_account_name=cash_acc.name,
            memo=memo or None,
        )

        if purchase_id:
            purchase = Purchase.query.filter_by(id=int(purchase_id), access_code_id=acc.id).first()
            if purchase:
                payment.purchase_id = purchase.id
                payment.supplier_name = purchase.supplier_name
                if amount >= float(purchase.total_amount or 0):
                    purchase.is_paid = True

        db.session.add(payment)
        db.session.flush()

        entry = _create_journal_for_ap_payment(payment)
        payment.journal_entry_id = entry.id

        db.session.commit()
        flash("Pembayaran hutang berhasil dicatat.", "success")
        return redirect(url_for("main.ap_payment_home"))

    payments = (
        APayment.query.filter_by(access_code_id=acc.id)
        .order_by(APayment.date.desc(), APayment.id.desc())
        .limit(20)
        .all()
    )
    return render_template(
        "ap_payment_home.html",
        purchases=purchases,
        cash_accounts=cash_accounts,
        payments=payments,
    )


@bp.route("/ap-payment/<int:payment_id>/edit", methods=["GET", "POST"])
def ap_payment_edit(payment_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    payment = APayment.query.filter_by(id=payment_id, access_code_id=acc.id).first_or_404()

    purchases = Purchase.query.filter_by(access_code_id=acc.id).order_by(Purchase.date.desc()).all()
    cash_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type == "Kas & Bank")
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        purchase_id = (request.form.get("purchase_id") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not cash_code or not amount_str:
            flash("Tanggal, akun kas, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.ap_payment_edit", payment_id=payment.id))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.ap_payment_edit", payment_id=payment.id))

        # rollback status pembelian lama
        if payment.purchase_id:
            old_purchase = Purchase.query.filter_by(id=payment.purchase_id, access_code_id=acc.id).first()
            if old_purchase:
                old_purchase.is_paid = False

        # hapus jurnal lama
        _delete_journal_entry_scoped(acc, payment.journal_entry_id)
        db.session.flush()

        # update payment
        payment.date = _parse_date(date_str)
        payment.amount = amount
        payment.memo = memo or None

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        if not cash_acc:
            flash("Akun kas/bank tidak valid.", "error")
            return redirect(url_for("main.ap_payment_edit", payment_id=payment.id))

        payment.cash_account_code = cash_acc.code
        payment.cash_account_name = cash_acc.name

        if purchase_id:
            purchase = Purchase.query.filter_by(id=int(purchase_id), access_code_id=acc.id).first()
            if purchase:
                payment.purchase_id = purchase.id
                payment.supplier_name = purchase.supplier_name
                if amount >= float(purchase.total_amount or 0):
                    purchase.is_paid = True
        else:
            payment.purchase_id = None
            payment.supplier_name = None

        entry = _create_journal_for_ap_payment(payment)
        payment.journal_entry_id = entry.id

        db.session.commit()
        flash("Pembayaran hutang berhasil diupdate.", "success")
        return redirect(url_for("main.ap_payment_home"))

    return render_template(
        "ap_payment_edit.html",
        payment=payment,
        purchases=purchases,
        cash_accounts=cash_accounts,
    )


@bp.post("/ap-payment/<int:payment_id>/delete")
def ap_payment_delete(payment_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    payment = APayment.query.filter_by(id=payment_id, access_code_id=acc.id).first_or_404()

    # rollback status hutang
    if payment.purchase_id:
        purchase = Purchase.query.filter_by(id=payment.purchase_id, access_code_id=acc.id).first()
        if purchase:
            purchase.is_paid = False

    _delete_journal_entry_scoped(acc, payment.journal_entry_id)

    db.session.delete(payment)
    db.session.commit()

    flash("Pembayaran hutang dihapus. Jurnal & status hutang dikembalikan.", "success")
    return redirect(url_for("main.ap_payment_home"))


# ============================================================
# Sales (SIMPLE) — pakai CashTransaction memo [SALE] (scoped)
# ============================================================
def _sale_memo(customer: str | None, note: str | None) -> str:
    customer = (customer or "").strip()
    note = (note or "").strip()
    parts = []
    if customer:
        parts.append(customer)
    if note:
        parts.append(note)
    suffix = " - ".join(parts) if parts else ""
    return "[SALE]" + (f" {suffix}" if suffix else "")


@bp.route("/sales", methods=["GET", "POST"])
def sales_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    debit_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["Kas & Bank", "Akun Piutang"]))
        .order_by(Account.code.asc())
        .all()
    )
    revenue_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["Pendapatan", "Pendapatan Lain"]))
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        customer = (request.form.get("customer_name") or "").strip()
        debit_code = (request.form.get("debit_account") or "").strip()
        credit_code = (request.form.get("revenue_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        note = (request.form.get("memo") or "").strip()

        if not date_str or not debit_code or not credit_code or not amount_str:
            flash("Tanggal, akun debit, akun pendapatan, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.sales_home"))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.sales_home"))

        debit_acc = Account.query.filter_by(access_code_id=acc.id, code=debit_code).first()
        credit_acc = Account.query.filter_by(access_code_id=acc.id, code=credit_code).first()
        if not debit_acc or not credit_acc:
            flash("Akun tidak valid.", "error")
            return redirect(url_for("main.sales_home"))

        tx = CashTransaction(
            access_code_id=acc.id,
            date=_parse_date(date_str),
            direction="in",
            cash_account_code=debit_acc.code,
            cash_account_name=debit_acc.name,
            counter_account_code=credit_acc.code,
            counter_account_name=credit_acc.name,
            amount=amount,
            memo=_sale_memo(customer, note),
        )
        db.session.add(tx)
        db.session.flush()

        entry = _create_journal_for_cash(tx)
        tx.journal_entry_id = entry.id

        db.session.commit()
        flash("Penjualan tersimpan & jurnal otomatis dibuat.", "success")
        return redirect(url_for("main.sales_home"))

    sales = (
        CashTransaction.query.filter_by(access_code_id=acc.id)
        .filter(CashTransaction.direction == "in")
        .filter(CashTransaction.memo.like("[SALE]%"))
        .order_by(CashTransaction.date.desc(), CashTransaction.id.desc())
        .limit(100)
        .all()
    )

    return render_template(
        "sales_home.html",
        debit_accounts=debit_accounts,
        revenue_accounts=revenue_accounts,
        sales=sales,
        today=datetime.utcnow().strftime("%Y-%m-%d"),
    )


@bp.route("/sales/<int:tx_id>/edit", methods=["GET", "POST"])
def sales_edit(tx_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    tx = CashTransaction.query.filter_by(id=tx_id, access_code_id=acc.id).first_or_404()

    if not (tx.direction == "in" and (tx.memo or "").startswith("[SALE]")):
        flash("Transaksi ini bukan penjualan.", "error")
        return redirect(url_for("main.sales_home"))

    debit_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["Kas & Bank", "Akun Piutang"]))
        .order_by(Account.code.asc())
        .all()
    )
    revenue_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["Pendapatan", "Pendapatan Lain"]))
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        customer = (request.form.get("customer_name") or "").strip()
        debit_code = (request.form.get("debit_account") or "").strip()
        credit_code = (request.form.get("revenue_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        note = (request.form.get("memo") or "").strip()

        if not date_str or not debit_code or not credit_code or not amount_str:
            flash("Tanggal, akun debit, akun pendapatan, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.sales_edit", tx_id=tx.id))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.sales_edit", tx_id=tx.id))

        debit_acc = Account.query.filter_by(access_code_id=acc.id, code=debit_code).first()
        credit_acc = Account.query.filter_by(access_code_id=acc.id, code=credit_code).first()
        if not debit_acc or not credit_acc:
            flash("Akun tidak valid.", "error")
            return redirect(url_for("main.sales_edit", tx_id=tx.id))

        # hapus jurnal lama
        if tx.journal_entry_id:
            old_entry = JournalEntry.query.filter_by(id=tx.journal_entry_id, access_code_id=acc.id).first()
            if old_entry:
                db.session.delete(old_entry)
                db.session.flush()

        # update tx
        tx.date = _parse_date(date_str)
        tx.cash_account_code = debit_acc.code
        tx.cash_account_name = debit_acc.name
        tx.counter_account_code = credit_acc.code
        tx.counter_account_name = credit_acc.name
        tx.amount = amount
        tx.memo = _sale_memo(customer, note)

        db.session.flush()

        new_entry = _create_journal_for_cash(tx)
        tx.journal_entry_id = new_entry.id

        db.session.commit()
        flash("Penjualan berhasil diupdate.", "success")
        return redirect(url_for("main.sales_home"))

    raw = (tx.memo or "").replace("[SALE]", "").strip()
    return render_template(
        "sales_edit.html",
        tx=tx,
        debit_accounts=debit_accounts,
        revenue_accounts=revenue_accounts,
        raw_memo=raw,
    )


@bp.post("/sales/<int:tx_id>/delete")
def sales_delete(tx_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    tx = CashTransaction.query.filter_by(id=tx_id, access_code_id=acc.id).first_or_404()

    if not (tx.direction == "in" and (tx.memo or "").startswith("[SALE]")):
        flash("Transaksi ini bukan penjualan.", "error")
        return redirect(url_for("main.sales_home"))

    if tx.journal_entry_id:
        entry = JournalEntry.query.filter_by(id=tx.journal_entry_id, access_code_id=acc.id).first()
        if entry:
            db.session.delete(entry)

    db.session.delete(tx)
    db.session.commit()

    flash("Penjualan dihapus.", "success")
    return redirect(url_for("main.sales_home"))


# ============================================================
# AR Payments (Invoice) — scoped (route kamu /ar/payments)
# ============================================================
@bp.route("/ar/payments", methods=["GET", "POST"])
def ar_payment_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    cash_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type == "Kas & Bank")
        .order_by(Account.code.asc())
        .all()
    )
    open_invoices = (
        SalesInvoice.query.filter_by(access_code_id=acc.id)
        .filter(SalesInvoice.status != "paid")
        .order_by(SalesInvoice.date.desc(), SalesInvoice.id.desc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        invoice_id = (request.form.get("invoice_id") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not invoice_id or not cash_code or not amount_str:
            flash("Tanggal, invoice, akun kas/bank, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.ar_payment_home"))

        inv = SalesInvoice.query.filter_by(id=int(invoice_id), access_code_id=acc.id).first()
        if not inv:
            flash("Invoice tidak ditemukan.", "error")
            return redirect(url_for("main.ar_payment_home"))

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        if not cash_acc:
            flash("Akun kas/bank tidak valid.", "error")
            return redirect(url_for("main.ar_payment_home"))

        try:
            amt = float(amount_str)
            if amt <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.ar_payment_home"))

        remaining = float(inv.total_amount or 0) - float(inv.paid_amount or 0)
        if amt > remaining:
            flash(f"Nominal melebihi sisa piutang (sisa: Rp {remaining:,.0f}).", "error")
            return redirect(url_for("main.ar_payment_home"))

        pay = ARPayment(
            access_code_id=acc.id,
            date=_parse_date(date_str),
            invoice_id=inv.id,
            invoice_no=inv.invoice_no,
            cash_account_code=cash_acc.code,
            cash_account_name=cash_acc.name,
            amount=amt,
            memo=memo or None,
        )
        db.session.add(pay)
        db.session.flush()

        entry = _create_journal_for_ar_payment(pay, inv)
        pay.journal_entry_id = entry.id

        inv.paid_amount = float(inv.paid_amount or 0) + amt
        if inv.paid_amount >= float(inv.total_amount or 0):
            inv.status = "paid"
            inv.paid_amount = float(inv.total_amount or 0)
        else:
            inv.status = "partial"

        db.session.commit()
        flash("Pembayaran piutang tersimpan & jurnal otomatis dibuat.", "success")
        return redirect(url_for("main.ar_payment_home"))

    payments = (
        ARPayment.query.filter_by(access_code_id=acc.id)
        .order_by(ARPayment.date.desc(), ARPayment.id.desc())
        .limit(50)
        .all()
    )
    return render_template(
        "ar_payment_home.html",
        payments=payments,
        cash_accounts=cash_accounts,
        open_invoices=open_invoices,
    )


# ============================================================
# Expenses (kas keluar ke akun beban) — scoped
# ============================================================
@bp.route("/expenses", methods=["GET", "POST"])
def expenses_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    cash_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type == "Kas & Bank")
        .order_by(Account.code.asc())
        .all()
    )
    expense_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["Beban", "Beban Lain"]))
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        exp_code = (request.form.get("expense_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not cash_code or not exp_code or not amount_str:
            flash("Tanggal, akun kas, akun beban, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.expenses_home"))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.expenses_home"))

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        exp_acc = Account.query.filter_by(access_code_id=acc.id, code=exp_code).first()
        if not cash_acc or not exp_acc:
            flash("Akun tidak valid.", "error")
            return redirect(url_for("main.expenses_home"))

        tx = CashTransaction(
            access_code_id=acc.id,
            date=_parse_date(date_str),
            direction="out",
            cash_account_code=cash_acc.code,
            cash_account_name=cash_acc.name,
            counter_account_code=exp_acc.code,
            counter_account_name=exp_acc.name,
            amount=amount,
            memo=memo or None,
        )
        db.session.add(tx)
        db.session.flush()

        entry = _create_journal_for_cash(tx)
        tx.journal_entry_id = entry.id

        db.session.commit()
        flash("Biaya operasional tersimpan & jurnal dibuat.", "success")
        return redirect(url_for("main.expenses_home"))

    txs = (
        CashTransaction.query.filter_by(access_code_id=acc.id, direction="out")
        .order_by(CashTransaction.date.desc(), CashTransaction.id.desc())
        .limit(50)
        .all()
    )
    return render_template("expenses_home.html", cash_accounts=cash_accounts, expense_accounts=expense_accounts, txs=txs)


@bp.route("/expenses/<int:tx_id>/edit", methods=["GET", "POST"])
def expense_edit(tx_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    tx = CashTransaction.query.filter_by(id=tx_id, access_code_id=acc.id).first_or_404()
    if tx.direction != "out":
        flash("Transaksi ini bukan transaksi biaya.", "error")
        return redirect(url_for("main.expenses_home"))

    cash_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type == "Kas & Bank")
        .order_by(Account.code.asc())
        .all()
    )
    expense_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["Beban", "Beban Lain"]))
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        exp_code = (request.form.get("expense_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not cash_code or not exp_code or not amount_str:
            flash("Tanggal, akun kas, akun beban, dan nominal wajib diisi.", "error")
            return redirect(url_for("main.expense_edit", tx_id=tx.id))

        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.expense_edit", tx_id=tx.id))

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        exp_acc = Account.query.filter_by(access_code_id=acc.id, code=exp_code).first()
        if not cash_acc or not exp_acc:
            flash("Akun tidak valid.", "error")
            return redirect(url_for("main.expense_edit", tx_id=tx.id))

        # update transaksi
        tx.date = _parse_date(date_str)
        tx.cash_account_code = cash_acc.code
        tx.cash_account_name = cash_acc.name
        tx.counter_account_code = exp_acc.code
        tx.counter_account_name = exp_acc.name
        tx.amount = amount
        tx.memo = memo or None

        # rebuild jurnal
        if getattr(tx, "journal_entry_id", None):
            old = JournalEntry.query.filter_by(id=tx.journal_entry_id, access_code_id=acc.id).first()
            if old:
                db.session.delete(old)
                db.session.flush()

        entry = _create_journal_for_cash(tx)
        tx.journal_entry_id = entry.id

        db.session.commit()
        flash("Transaksi biaya berhasil diupdate.", "success")
        return redirect(url_for("main.expenses_home"))

    return render_template(
        "expense_edit.html",
        tx=tx,
        cash_accounts=cash_accounts,
        expense_accounts=expense_accounts,
    )


@bp.post("/expenses/<int:tx_id>/delete")
def expense_delete(tx_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    tx = CashTransaction.query.filter_by(id=tx_id, access_code_id=acc.id).first_or_404()
    if tx.direction != "out":
        flash("Transaksi ini bukan transaksi biaya.", "error")
        return redirect(url_for("main.expenses_home"))

    if getattr(tx, "journal_entry_id", None):
        old = JournalEntry.query.filter_by(id=tx.journal_entry_id, access_code_id=acc.id).first()
        if old:
            db.session.delete(old)

    db.session.delete(tx)
    db.session.commit()
    flash("Transaksi biaya berhasil dihapus.", "success")
    return redirect(url_for("main.expenses_home"))


# ============================================================
# Stock Usage — scoped
# ============================================================
@bp.route("/stock-usage", methods=["GET", "POST"])
def stock_usage_home():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    items = Item.query.filter_by(access_code_id=acc.id).order_by(Item.name.asc()).all()
    hpp_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["HPP", "Beban"]))
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        item_id = (request.form.get("item_id") or "").strip()
        qty_str = (request.form.get("qty") or "").strip()
        hpp_code = (request.form.get("hpp_account") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not item_id or not qty_str or not hpp_code:
            flash("Tanggal, bahan, qty, dan akun HPP wajib diisi.", "error")
            return redirect(url_for("main.stock_usage_home"))

        try:
            qty = float(qty_str)
            if qty <= 0:
                raise ValueError()
        except ValueError:
            flash("Qty harus angka > 0.", "error")
            return redirect(url_for("main.stock_usage_home"))

        item = Item.query.filter_by(id=int(item_id), access_code_id=acc.id).first()
        if not item:
            flash("Bahan tidak valid.", "error")
            return redirect(url_for("main.stock_usage_home"))

        if float(item.stock_qty or 0) < qty:
            flash(f"Stok tidak cukup. Stok saat ini: {item.stock_qty:g} {item.unit}.", "error")
            return redirect(url_for("main.stock_usage_home"))

        hpp_acc = Account.query.filter_by(access_code_id=acc.id, code=hpp_code).first()
        if not hpp_acc:
            flash("Akun HPP tidak valid.", "error")
            return redirect(url_for("main.stock_usage_home"))

        unit_cost = float(item.avg_cost or 0)
        total_cost = qty * unit_cost

        u = StockUsage(
            access_code_id=acc.id,
            date=_parse_date(date_str),
            item_id=item.id,
            item_name=item.name,
            qty=qty,
            unit_cost=unit_cost,
            total_cost=total_cost,
            hpp_account_code=hpp_acc.code,
            hpp_account_name=hpp_acc.name,
            memo=memo or None,
        )
        db.session.add(u)
        db.session.flush()

        item.stock_qty = float(item.stock_qty or 0) - qty

        entry = _create_journal_for_stock_usage(u)
        u.journal_entry_id = entry.id

        db.session.commit()
        flash("Pemakaian stok tersimpan, persediaan berkurang, jurnal dibuat.", "success")
        return redirect(url_for("main.stock_usage_home"))

    usages = (
        StockUsage.query.filter_by(access_code_id=acc.id)
        .order_by(StockUsage.date.desc(), StockUsage.id.desc())
        .limit(50)
        .all()
    )
    return render_template("stock_usage_home.html", items=items, hpp_accounts=hpp_accounts, usages=usages)


@bp.route("/stock-usage/<int:usage_id>/edit", methods=["GET", "POST"])
def stock_usage_edit(usage_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    usage = StockUsage.query.filter_by(id=usage_id, access_code_id=acc.id).first_or_404()

    items = Item.query.filter_by(access_code_id=acc.id).order_by(Item.name.asc()).all()
    hpp_accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .filter(Account.type.in_(["HPP", "Beban"]))
        .order_by(Account.code.asc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        item_id_str = (request.form.get("item_id") or "").strip()
        qty_str = (request.form.get("qty") or "").strip()
        hpp_code = (request.form.get("hpp_account") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not item_id_str or not qty_str or not hpp_code:
            flash("Tanggal, bahan, qty, dan akun HPP wajib diisi.", "error")
            return redirect(url_for("main.stock_usage_edit", usage_id=usage.id))

        try:
            new_qty = float(qty_str)
            if new_qty <= 0:
                raise ValueError()
        except ValueError:
            flash("Qty harus angka > 0.", "error")
            return redirect(url_for("main.stock_usage_edit", usage_id=usage.id))

        new_item = Item.query.filter_by(id=int(item_id_str), access_code_id=acc.id).first()
        if not new_item:
            flash("Bahan tidak valid.", "error")
            return redirect(url_for("main.stock_usage_edit", usage_id=usage.id))

        hpp_acc = Account.query.filter_by(access_code_id=acc.id, code=hpp_code).first()
        if not hpp_acc:
            flash("Akun HPP tidak valid.", "error")
            return redirect(url_for("main.stock_usage_edit", usage_id=usage.id))

        # 1) balikin stok dari pemakaian lama
        old_item = Item.query.filter_by(id=usage.item_id, access_code_id=acc.id).first()
        old_qty = float(usage.qty or 0)
        if old_item:
            old_item.stock_qty = float(old_item.stock_qty or 0) + old_qty

        # 2) cek stok cukup untuk pemakaian baru (setelah rollback)
        if float(new_item.stock_qty or 0) < new_qty:
            flash(
                f"Stok tidak cukup setelah penyesuaian. Stok tersedia: {float(new_item.stock_qty or 0):g} {new_item.unit}.",
                "error",
            )
            db.session.rollback()
            return redirect(url_for("main.stock_usage_edit", usage_id=usage.id))

        # 3) apply pemakaian baru
        unit_cost = float(new_item.avg_cost or 0)
        total_cost = new_qty * unit_cost
        new_item.stock_qty = float(new_item.stock_qty or 0) - new_qty

        usage.date = _parse_date(date_str)
        usage.item_id = new_item.id
        usage.item_name = new_item.name
        usage.qty = new_qty
        usage.unit_cost = unit_cost
        usage.total_cost = total_cost
        usage.hpp_account_code = hpp_acc.code
        usage.hpp_account_name = hpp_acc.name
        usage.memo = memo or None

        # 4) rebuild jurnal
        if getattr(usage, "journal_entry_id", None):
            old_entry = JournalEntry.query.filter_by(id=usage.journal_entry_id, access_code_id=acc.id).first()
            if old_entry:
                db.session.delete(old_entry)
                db.session.flush()

        entry = _create_journal_for_stock_usage(usage)
        usage.journal_entry_id = entry.id

        db.session.commit()
        flash("Pemakaian stok berhasil diupdate.", "success")
        return redirect(url_for("main.stock_usage_home"))

    return render_template(
        "stock_usage_edit.html",
        usage=usage,
        items=items,
        hpp_accounts=hpp_accounts,
    )


@bp.post("/stock-usage/<int:usage_id>/delete")
def stock_usage_delete(usage_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    usage = StockUsage.query.filter_by(id=usage_id, access_code_id=acc.id).first_or_404()

    # balikin stok
    item = Item.query.filter_by(id=usage.item_id, access_code_id=acc.id).first()
    if item:
        item.stock_qty = float(item.stock_qty or 0) + float(usage.qty or 0)

    # hapus jurnal terkait
    if getattr(usage, "journal_entry_id", None):
        old_entry = JournalEntry.query.filter_by(id=usage.journal_entry_id, access_code_id=acc.id).first()
        if old_entry:
            db.session.delete(old_entry)

    db.session.delete(usage)
    db.session.commit()
    flash("Pemakaian stok berhasil dihapus (stok & jurnal dikembalikan).", "success")
    return redirect(url_for("main.stock_usage_home"))

# ============================================================
# PART 4/4 — REPORTS + EXPORTS + AUDIT + REBUILD HELPERS (scoped)
#   - Report Ledger (web)
#   - Excel helpers + export ledger.xlsx + ledger-all.xlsx
#   - Report Profit/Loss (web) + export PDF
#   - Report Balance Sheet (web) + export PDF
#   - Export Sales Invoice PDF
#   - Admin audit unbalanced
#   - Rebuild helpers (inventory + paid flags + journals)
#   - AR Payment edit/delete (yang versi bawah file kamu)
# ============================================================

from sqlalchemy import func, and_

# ============================================================
# REPORT: Buku Besar (filter tanggal) — scoped
# ============================================================
@bp.get("/reports/ledger")
def report_ledger():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    accounts = (
        Account.query.filter_by(access_code_id=acc.id)
        .order_by(Account.code.asc())
        .all()
    )
    selected_code = (request.args.get("account") or "").strip()

    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()

    from_date = _parse_date(from_str) if from_str else None
    to_date = _parse_date(to_str) if to_str else None

    lines = []
    balance = 0.0

    if selected_code:
        # pastikan akun ini milik acc
        a = Account.query.filter_by(access_code_id=acc.id, code=selected_code).first()
        if not a:
            flash("Akun tidak ditemukan untuk dapur ini.", "error")
            return redirect(url_for("main.report_ledger"))

        q = _jl_base_query(from_date, (to_date + timedelta(days=1)) if to_date else None).filter(
            JournalLine.access_code_id == acc.id,
            JournalLine.account_code == selected_code,
        )

        lines = q.order_by(JournalEntry.date.asc(), JournalLine.id.asc()).all()
        balance = sum((ln.debit or 0) - (ln.credit or 0) for ln in lines)

    return render_template(
        "report_ledger.html",
        accounts=accounts,
        selected_code=selected_code or None,
        lines=lines,
        balance=balance,
        from_date=from_str,
        to_date=to_str,
    )


# =========================
# Excel Helpers
# =========================
_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _autosize_columns(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))

def _style_header_row(ws, row_idx=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = _BORDER
    ws.row_dimensions[row_idx].height = 20

def _style_table_cells(ws, start_row, end_row, start_col, end_col):
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="top")
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            if c in (3, 4, 5):
                cell.alignment = align_right
            else:
                cell.alignment = align_left

def _fmt_idr_excel(cell):
    cell.number_format = "#,##0"

def _get_entry_date_and_memo(line: JournalLine):
    je = None
    if hasattr(line, "entry") and line.entry is not None:
        je = line.entry
    else:
        if hasattr(line, "entry_id"):
            je = JournalEntry.query.get(line.entry_id)
        elif hasattr(line, "journal_entry_id"):
            je = JournalEntry.query.get(line.journal_entry_id)
    if not je:
        return None, "-"
    return je.date, (je.memo or "-")


# =========================
# EXPORT: Buku Besar ke Excel (Per Akun) — scoped
# URL: /export/ledger.xlsx?account=10011&from=2026-02-01&to=2026-02-02
# =========================
@bp.get("/export/ledger.xlsx")
def export_ledger_xlsx():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    code = (request.args.get("account") or "").strip()
    if not code:
        flash("Pilih akun dulu untuk export Buku Besar.", "error")
        return redirect(url_for("main.report_ledger"))

    from_dt, to_dt_excl, from_str, to_str = _get_date_range_args()

    account = Account.query.filter_by(access_code_id=acc.id, code=code).first()
    if not account:
        flash("Akun tidak ditemukan untuk dapur ini.", "error")
        return redirect(url_for("main.report_ledger"))

    q = _jl_base_query(from_dt, to_dt_excl).filter(
        JournalLine.access_code_id == acc.id,
        JournalLine.account_code == code,
    )
    q = q.order_by(JournalEntry.date.asc(), JournalLine.id.asc())
    lines = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Besar"

    ws["A1"] = "Buku Besar"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Akun: {code} - {account.name}".strip(" -")
    ws["A3"] = f"Dapur: {acc.dapur_name or 'Dapur MBG'}"

    periode = "Periode: "
    if from_str and to_str:
        periode += f"{from_str} s/d {to_str}"
    elif from_str and not to_str:
        periode += f"mulai {from_str}"
    elif (not from_str) and to_str:
        periode += f"sampai {to_str}"
    else:
        periode += "Seluruh Periode"
    ws["A4"] = periode

    start_row = 6
    ws.append([""] * 5)  # row 5
    ws.append(["Tanggal", "Keterangan", "Debit", "Kredit", "Saldo Berjalan"])  # row 6
    _style_header_row(ws, start_row)

    saldo = 0.0
    r = start_row + 1
    for ln in lines:
        dt, memo = _get_entry_date_and_memo(ln)
        debit = float(ln.debit or 0)
        credit = float(ln.credit or 0)
        saldo += (debit - credit)

        ws.cell(row=r, column=1, value=dt.date().isoformat() if dt else "-")
        ws.cell(row=r, column=2, value=memo)
        ws.cell(row=r, column=3, value=debit)
        ws.cell(row=r, column=4, value=credit)
        ws.cell(row=r, column=5, value=saldo)

        _fmt_idr_excel(ws.cell(row=r, column=3))
        _fmt_idr_excel(ws.cell(row=r, column=4))
        _fmt_idr_excel(ws.cell(row=r, column=5))
        r += 1

    if r > start_row + 1:
        _style_table_cells(ws, start_row + 1, r - 1, 1, 5)

    total_row = r + 1
    ws.cell(row=total_row, column=2, value="SALDO AKHIR").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=saldo).font = Font(bold=True)
    _fmt_idr_excel(ws.cell(row=total_row, column=5))

    _autosize_columns(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"buku_besar_{code}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================
# EXPORT: Buku Besar ke Excel (Semua Akun, per sheet) — scoped
# URL: /export/ledger-all.xlsx?from=2026-02-01&to=2026-02-02
# =========================
@bp.get("/export/ledger-all.xlsx")
def export_ledger_all_xlsx():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    from_dt, to_dt_excl, from_str, to_str = _get_date_range_args()

    periode = "Periode: "
    if from_str and to_str:
        periode += f"{from_str} s/d {to_str}"
    elif from_str and not to_str:
        periode += f"mulai {from_str}"
    elif (not from_str) and to_str:
        periode += f"sampai {to_str}"
    else:
        periode += "Seluruh Periode"

    accounts = Account.query.filter_by(access_code_id=acc.id).order_by(Account.code.asc()).all()

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Ringkasan"

    ws_sum["A1"] = "Buku Besar - Semua Akun"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Dapur: {acc.dapur_name or 'Dapur MBG'}"
    ws_sum["A3"] = periode

    ws_sum.append([])
    ws_sum.append(["Kode", "Nama Akun", "Saldo (Debit - Kredit)"])
    header_row = ws_sum.max_row
    _style_header_row(ws_sum, header_row)

    ringkasan_start = ws_sum.max_row + 1
    ringkasan_row = ringkasan_start

    for a in accounts:
        q = _jl_base_query(from_dt, to_dt_excl).filter(
            JournalLine.access_code_id == acc.id,
            JournalLine.account_code == a.code,
        )
        q = q.order_by(JournalEntry.date.asc(), JournalLine.id.asc())
        lines = q.all()
        if not lines:
            continue

        sheet_name = f"{a.code}"
        if sheet_name in wb.sheetnames:
            sheet_name = f"{a.code}_{(a.name or '')[:10]}"
        ws = wb.create_sheet(title=sheet_name[:31])

        ws["A1"] = "Buku Besar"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Akun: {a.code} - {a.name}"
        ws["A3"] = f"Dapur: {acc.dapur_name or 'Dapur MBG'}"
        ws["A4"] = periode

        start_row = 6
        ws.append([""] * 5)
        ws.append(["Tanggal", "Keterangan", "Debit", "Kredit", "Saldo Berjalan"])
        _style_header_row(ws, start_row)

        saldo = 0.0
        r = start_row + 1
        for ln in lines:
            dt, memo = _get_entry_date_and_memo(ln)
            debit = float(ln.debit or 0)
            credit = float(ln.credit or 0)
            saldo += (debit - credit)

            ws.cell(row=r, column=1, value=dt.date().isoformat() if dt else "-")
            ws.cell(row=r, column=2, value=memo)
            ws.cell(row=r, column=3, value=debit)
            ws.cell(row=r, column=4, value=credit)
            ws.cell(row=r, column=5, value=saldo)

            _fmt_idr_excel(ws.cell(row=r, column=3))
            _fmt_idr_excel(ws.cell(row=r, column=4))
            _fmt_idr_excel(ws.cell(row=r, column=5))
            r += 1

        if r > start_row + 1:
            _style_table_cells(ws, start_row + 1, r - 1, 1, 5)

        ws.cell(row=r + 1, column=2, value="SALDO AKHIR").font = Font(bold=True)
        ws.cell(row=r + 1, column=5, value=saldo).font = Font(bold=True)
        _fmt_idr_excel(ws.cell(row=r + 1, column=5))

        _autosize_columns(ws)

        # ringkasan
        ws_sum.cell(row=ringkasan_row, column=1, value=a.code)
        ws_sum.cell(row=ringkasan_row, column=2, value=a.name)
        ws_sum.cell(row=ringkasan_row, column=3, value=saldo)
        _fmt_idr_excel(ws_sum.cell(row=ringkasan_row, column=3))
        ringkasan_row += 1

    if ringkasan_row > ringkasan_start:
        _style_table_cells(ws_sum, ringkasan_start, ringkasan_row - 1, 1, 3)

    _autosize_columns(ws_sum)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="buku_besar_semua_akun.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# REPORT: Laba Rugi (filter tanggal, struktur standar) — scoped
# ============================================================
@bp.get("/reports/profit-loss")
def report_profit_loss():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    dfrom, dto = _get_date_range_from_request()

    rev_main = Account.query.filter_by(access_code_id=acc.id, type="Pendapatan").order_by(Account.code.asc()).all()
    hpp_accounts = Account.query.filter_by(access_code_id=acc.id, type="HPP").order_by(Account.code.asc()).all()
    op_exp = Account.query.filter_by(access_code_id=acc.id, type="Beban").order_by(Account.code.asc()).all()
    rev_other = Account.query.filter_by(access_code_id=acc.id, type="Pendapatan Lain").order_by(Account.code.asc()).all()
    exp_other = Account.query.filter_by(access_code_id=acc.id, type="Beban Lain").order_by(Account.code.asc()).all()

    def amt_revenue(a):
        return -_account_balance_range(a.code, dfrom, dto)

    def amt_expense(a):
        return _account_balance_range(a.code, dfrom, dto)

    rev_main_data, total_rev_main = [], 0.0
    for a in rev_main:
        amt = float(amt_revenue(a))
        if amt != 0:
            rev_main_data.append((a, amt))
            total_rev_main += amt

    hpp_data, total_hpp = [], 0.0
    for a in hpp_accounts:
        amt = float(amt_expense(a))
        if amt != 0:
            hpp_data.append((a, amt))
            total_hpp += amt

    gross_profit = total_rev_main - total_hpp

    op_exp_data, total_op_exp = [], 0.0
    for a in op_exp:
        amt = float(amt_expense(a))
        if amt != 0:
            op_exp_data.append((a, amt))
            total_op_exp += amt

    operating_profit = gross_profit - total_op_exp

    rev_other_data, total_rev_other = [], 0.0
    for a in rev_other:
        amt = float(amt_revenue(a))
        if amt != 0:
            rev_other_data.append((a, amt))
            total_rev_other += amt

    exp_other_data, total_exp_other = [], 0.0
    for a in exp_other:
        amt = float(amt_expense(a))
        if amt != 0:
            exp_other_data.append((a, amt))
            total_exp_other += amt

    net_profit = operating_profit + total_rev_other - total_exp_other

    return render_template(
        "report_profit_loss.html",
        rev_main_data=rev_main_data,
        hpp_data=hpp_data,
        op_exp_data=op_exp_data,
        rev_other_data=rev_other_data,
        exp_other_data=exp_other_data,
        total_rev_main=total_rev_main,
        total_hpp=total_hpp,
        gross_profit=gross_profit,
        total_op_exp=total_op_exp,
        operating_profit=operating_profit,
        total_rev_other=total_rev_other,
        total_exp_other=total_exp_other,
        net_profit=net_profit,
        dfrom=dfrom.strftime("%Y-%m-%d"),
        dto=dto.strftime("%Y-%m-%d"),
    )


# ============================================================
# REPORT: Neraca (filter tanggal + laba tahun berjalan) — scoped
# ============================================================
@bp.get("/reports/balance-sheet")
def report_balance_sheet():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    as_of_str = (request.args.get("as_of") or "").strip()
    as_of_date = _parse_ymd(as_of_str) or datetime.utcnow().date()
    to_dt_excl = datetime.combine(as_of_date + timedelta(days=1), datetime.min.time())

    def bal_upto(code: str) -> float:
        fk = _jl_entry_fk()
        q = (
            JournalLine.query.join(JournalEntry, fk == JournalEntry.id)
            .filter(
                JournalLine.access_code_id == acc.id,
                JournalLine.account_code == code,
                JournalEntry.access_code_id == acc.id,
                JournalEntry.date < to_dt_excl,
            )
        )
        debit = q.with_entities(db.func.coalesce(db.func.sum(JournalLine.debit), 0.0)).scalar() or 0.0
        credit = q.with_entities(db.func.coalesce(db.func.sum(JournalLine.credit), 0.0)).scalar() or 0.0
        return float(debit) - float(credit)

    assets = Account.query.filter_by(access_code_id=acc.id).filter(
        Account.type.in_([
            "Kas & Bank",
            "Akun Piutang",
            "Aktiva Lancar Lain",
            "Persediaan",
            "Aktiva Tetap",
            "Akum. Peny.",
        ])
    ).order_by(Account.code.asc()).all()

    liabilities = Account.query.filter_by(access_code_id=acc.id).filter(
        Account.type.in_(["Akun Hutang", "Hutang Lancar Lain", "Hutang Jk. Panjang"])
    ).order_by(Account.code.asc()).all()

    equities = Account.query.filter_by(access_code_id=acc.id).filter(
        Account.type == "Ekuitas"
    ).order_by(Account.code.asc()).all()

    asset_data, liab_data, eq_data = [], [], []
    total_assets = total_liab = total_eq = 0.0

    for a in assets:
        amt = float(bal_upto(a.code))
        if amt != 0:
            asset_data.append((a, amt))
            total_assets += amt

    for a in liabilities:
        amt = -float(bal_upto(a.code))
        if amt != 0:
            liab_data.append((a, amt))
            total_liab += amt

    for a in equities:
        amt = -float(bal_upto(a.code))
        if amt != 0:
            eq_data.append((a, amt))
            total_eq += amt

    # NET PROFIT sampai as_of (scoped)
    rev_accounts = Account.query.filter_by(access_code_id=acc.id, type="Pendapatan").all()
    rev_other_accounts = Account.query.filter_by(access_code_id=acc.id, type="Pendapatan Lain").all()
    hpp_accounts = Account.query.filter_by(access_code_id=acc.id, type="HPP").all()
    exp_accounts = Account.query.filter_by(access_code_id=acc.id, type="Beban").all()
    exp_other_accounts = Account.query.filter_by(access_code_id=acc.id, type="Beban Lain").all()

    sum_rev = sum(float(bal_upto(a.code)) for a in rev_accounts
    sum_rev = sum(float(bal_upto(a.code)) for a in rev_accounts)
    sum_rev_other = sum(float(bal_upto(a.code)) for a in rev_other_accounts)
    sum_hpp = sum(float(bal_upto(a.code)) for a in hpp_accounts)
    sum_exp = sum(float(bal_upto(a.code)) for a in exp_accounts)
    sum_exp_other = sum(float(bal_upto(a.code)) for a in exp_other_accounts)

    # pendapatan = kredit → negatif, jadi dibalik
    net_profit = (-sum_rev - sum_rev_other) - (sum_hpp + sum_exp + sum_exp_other)

    eq_data.append(("Laba (Rugi) Berjalan", net_profit))
    total_eq += net_profit

    return render_template(
        "report_balance_sheet.html",
        as_of=as_of_date.strftime("%Y-%m-%d"),
        asset_data=asset_data,
        liab_data=liab_data,
        eq_data=eq_data,
        total_assets=total_assets,
        total_liab=total_liab,
        total_eq=total_eq,
    )


# ============================================================
# EXPORT PDF: Laba Rugi
# ============================================================
@bp.get("/reports/profit-loss.pdf")
def export_profit_loss_pdf():
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    dfrom, dto = _get_date_range_from_request()

    buf = BytesIO()
    c, doc = pdf_doc(buf)

    header_block(
        c,
        title="Laporan Laba Rugi",
        subtitle=f"{dfrom.strftime('%d %b %Y')} s/d {dto.strftime('%d %b %Y')}",
        right=f"Dapur: {acc.dapur_name or '-'}",
    )

    y = 740
    y = section_title(c, "Pendapatan", y)
    total_rev = 0
    for a in Account.query.filter_by(access_code_id=acc.id, type="Pendapatan"):
        amt = -_account_balance_range(a.code, dfrom, dto)
        if amt != 0:
            table_2col(c, a.name, fmt_idr(amt), y)
            y -= 16
            total_rev += amt

    y -= 8
    table_2col(c, "Total Pendapatan", fmt_idr(total_rev), y, bold=True)
    y -= 24

    y = section_title(c, "HPP", y)
    total_hpp = 0
    for a in Account.query.filter_by(access_code_id=acc.id, type="HPP"):
        amt = _account_balance_range(a.code, dfrom, dto)
        if amt != 0:
            table_2col(c, a.name, fmt_idr(amt), y)
            y -= 16
            total_hpp += amt

    y -= 8
    table_2col(c, "Total HPP", fmt_idr(total_hpp), y, bold=True)
    y -= 24

    laba_kotor = total_rev - total_hpp
    table_2col(c, "LABA KOTOR", fmt_idr(laba_kotor), y, bold=True)
    y -= 32

    y = section_title(c, "Beban Operasional", y)
    total_exp = 0
    for a in Account.query.filter_by(access_code_id=acc.id, type="Beban"):
        amt = _account_balance_range(a.code, dfrom, dto)
        if amt != 0:
            table_2col(c, a.name, fmt_idr(amt), y)
            y -= 16
            total_exp += amt

    y -= 8
    table_2col(c, "Total Beban", fmt_idr(total_exp), y, bold=True)
    y -= 24

    laba_usaha = laba_kotor - total_exp
    table_2col(c, "LABA USAHA", fmt_idr(laba_usaha), y, bold=True)
    y -= 32

    footer_canvas(c)
    c.showPage()
    c.save()

    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="laba_rugi.pdf",
        mimetype="application/pdf",
    )


# ============================================================
# ADMIN AUDIT: jurnal tidak balance
# ============================================================
@bp.get("/admin/audit/unbalanced")
def admin_audit_unbalanced():
    guard = _require_admin()
    if guard:
        return guard

    rows = []
    entries = JournalEntry.query.order_by(JournalEntry.date.desc()).all()
    for e in entries:
        debit = sum(l.debit or 0 for l in e.lines)
        credit = sum(l.credit or 0 for l in e.lines)
        if round(debit - credit, 2) != 0:
            rows.append((e, debit, credit))

    return render_template("admin_audit_unbalanced.html", rows=rows)


# ============================================================
# REBUILD: inventory + paid flags + rebuild journals (ADMIN)
# ============================================================
@bp.post("/admin/rebuild/everything")
def admin_rebuild_everything():
    guard = _require_admin()
    if guard:
        return guard

    try:
        _rebuild_everything()
        flash("Rebuild berhasil: inventory, paid flags, dan seluruh jurnal sudah dibuat ulang.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Gagal rebuild: {e}", "error")

    return redirect(url_for("main.dashboard"))


# ============================================================
# EDIT / DELETE + REBUILD (STOK + JURNAL) — scoped helpers
# ============================================================

def _delete_journal_entry(entry_id: int | None, acc_id: int | None = None):
    """
    Hapus JournalLine + JournalEntry untuk entry_id.
    Kalau acc_id diberikan, pastikan entry memang milik dapur tersebut.
    """
    if not entry_id:
        return

    q_entry = JournalEntry.query.filter_by(id=entry_id)
    if acc_id is not None:
        q_entry = q_entry.filter_by(access_code_id=acc_id)

    entry = q_entry.first()
    if not entry:
        return

    # relationship sudah cascade? tetap aman manual delete lines by FK
    JournalLine.query.filter_by(entry_id=entry_id, access_code_id=entry.access_code_id).delete()
    db.session.delete(entry)


def _recalc_purchase_paid_flags(acc_id: int):
    """
    Set purchase.is_paid berdasarkan total pembayaran APayment per purchase (scoped).
    """
    purchases = Purchase.query.filter_by(access_code_id=acc_id).all()
    for p in purchases:
        total_paid = (
            db.session.query(db.func.coalesce(db.func.sum(APayment.amount), 0.0))
            .filter(APayment.access_code_id == acc_id, APayment.purchase_id == p.id)
            .scalar()
            or 0.0
        )
        total = float(p.total_amount or 0)
        p.is_paid = bool(total_paid >= total and total > 0)


def _recalc_invoice_paid_fields(acc_id: int):
    """
    Set invoice.paid_amount & status berdasarkan total pembayaran ARPayment per invoice (scoped).
    """
    invoices = SalesInvoice.query.filter_by(access_code_id=acc_id).all()
    for inv in invoices:
        total_paid = (
            db.session.query(db.func.coalesce(db.func.sum(ARPayment.amount), 0.0))
            .filter(ARPayment.access_code_id == acc_id, ARPayment.invoice_id == inv.id)
            .scalar()
            or 0.0
        )
        inv.paid_amount = float(total_paid)
        total = float(inv.total_amount or 0)

        if total <= 0:
            inv.status = "unpaid"
        elif inv.paid_amount <= 0:
            inv.status = "unpaid"
        elif inv.paid_amount >= total:
            inv.status = "paid"
            inv.paid_amount = total  # clamp
        else:
            inv.status = "partial"


def _rebuild_inventory(acc_id: int):
    """
    Rebuild stok & avg_cost semua item dari histori (scoped):
      + PurchaseItem (masuk)
      - StockUsage (keluar)
    Moving average sesuai logika kamu.
    """
    items = Item.query.filter_by(access_code_id=acc_id).all()
    for it in items:
        it.stock_qty = 0.0
        it.avg_cost = 0.0

    purchase_rows = (
        db.session.query(PurchaseItem, Purchase)
        .join(Purchase, PurchaseItem.purchase_id == Purchase.id)
        .filter(PurchaseItem.access_code_id == acc_id, Purchase.access_code_id == acc_id)
        .order_by(Purchase.date.asc(), Purchase.id.asc(), PurchaseItem.id.asc())
        .all()
    )

    usage_rows = (
        StockUsage.query
        .filter_by(access_code_id=acc_id)
        .order_by(StockUsage.date.asc(), StockUsage.id.asc())
        .all()
    )

    events = []
    for pi, p in purchase_rows:
        events.append((p.date, 0, "purchase", pi))
    for u in usage_rows:
        events.append((u.date, 1, "usage", u))
    events.sort(key=lambda x: (x[0] or datetime.min, x[1]))

    item_map = {it.id: it for it in items}

    for _, _, etype, obj in events:
        if etype == "purchase":
            pi: PurchaseItem = obj
            it = item_map.get(pi.item_id)
            if not it:
                continue
            qty = float(pi.qty or 0)
            price = float(pi.price or 0)
            if qty <= 0:
                continue

            total_cost_existing = float(it.stock_qty or 0) * float(it.avg_cost or 0)
            total_cost_new = qty * price
            new_qty = float(it.stock_qty or 0) + qty
            it.avg_cost = (total_cost_existing + total_cost_new) / new_qty if new_qty > 0 else 0.0
            it.stock_qty = new_qty

        elif etype == "usage":
            u: StockUsage = obj
            it = item_map.get(u.item_id)
            if not it:
                continue
            qty = float(u.qty or 0)
            if qty <= 0:
                continue

            it.stock_qty = float(it.stock_qty or 0) - qty
            if it.stock_qty < 0:
                it.stock_qty = 0.0


def _rebuild_all_journals(acc_id: int):
    """
    Hapus semua journal entries/lines milik access_code_id ini lalu buat ulang
    berdasarkan seluruh transaksi (scoped).
    """
    # delete lines then entries (scoped)
    JournalLine.query.filter_by(access_code_id=acc_id).delete()
    JournalEntry.query.filter_by(access_code_id=acc_id).delete()

    # reset FK jurnal di transaksi (scoped)
    CashTransaction.query.filter_by(access_code_id=acc_id).update({CashTransaction.journal_entry_id: None})
    Purchase.query.filter_by(access_code_id=acc_id).update({Purchase.journal_entry_id: None})
    APayment.query.filter_by(access_code_id=acc_id).update({APayment.journal_entry_id: None})
    StockUsage.query.filter_by(access_code_id=acc_id).update({StockUsage.journal_entry_id: None})
    SalesInvoice.query.filter_by(access_code_id=acc_id).update({SalesInvoice.journal_entry_id: None})
    ARPayment.query.filter_by(access_code_id=acc_id).update({ARPayment.journal_entry_id: None})

    db.session.flush()

    # 1) CashTransaction
    txs = (
        CashTransaction.query.filter_by(access_code_id=acc_id)
        .order_by(CashTransaction.date.asc(), CashTransaction.id.asc())
        .all()
    )
    for tx in txs:
        entry = _create_journal_for_cash(tx)
        tx.journal_entry_id = entry.id

    # 2) Purchase
    purchases = (
        Purchase.query.filter_by(access_code_id=acc_id)
        .order_by(Purchase.date.asc(), Purchase.id.asc())
        .all()
    )
    for p in purchases:
        entry = _create_journal_for_purchase(p)
        p.journal_entry_id = entry.id

    # 3) AP Payment
    pays = (
        APayment.query.filter_by(access_code_id=acc_id)
        .order_by(APayment.date.asc(), APayment.id.asc())
        .all()
    )
    for pay in pays:
        entry = _create_journal_for_ap_payment(pay)
        pay.journal_entry_id = entry.id

    # 4) Stock Usage
    usages = (
        StockUsage.query.filter_by(access_code_id=acc_id)
        .order_by(StockUsage.date.asc(), StockUsage.id.asc())
        .all()
    )
    for u in usages:
        entry = _create_journal_for_stock_usage(u)
        u.journal_entry_id = entry.id

    # 5) Sales Invoice
    invoices = (
        SalesInvoice.query.filter_by(access_code_id=acc_id)
        .order_by(SalesInvoice.date.asc(), SalesInvoice.id.asc())
        .all()
    )
    for inv in invoices:
        entry = _create_journal_for_invoice(inv)
        inv.journal_entry_id = entry.id

    # 6) AR Payment
    arps = (
        ARPayment.query.filter_by(access_code_id=acc_id)
        .order_by(ARPayment.date.asc(), ARPayment.id.asc())
        .all()
    )
    for p in arps:
        inv = SalesInvoice.query.filter_by(access_code_id=acc_id, id=p.invoice_id).first()
        if not inv:
            continue
        entry = _create_journal_for_ar_payment(p, inv)
        p.journal_entry_id = entry.id


def _rebuild_everything():
    """
    Scoped rebuild: hanya untuk access_code yang sedang login.
    Dipanggil setelah edit/hapus transaksi pada dapur tersebut.
    """
    acc = _require_access()
    if not acc:
        raise Exception("Tidak ada akses aktif.")

    _rebuild_inventory(acc.id)
    _recalc_purchase_paid_flags(acc.id)
    _recalc_invoice_paid_fields(acc.id)
    _rebuild_all_journals(acc.id)
    db.session.commit()


# ============================================================
# AR PAYMENT - EDIT / DELETE (yang versi bawah file kamu) — scoped
# ============================================================
@bp.route("/ar/payments/<int:pay_id>/edit", methods=["GET", "POST"])
def ar_payment_edit(pay_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    pay = ARPayment.query.filter_by(access_code_id=acc.id, id=pay_id).first_or_404()

    cash_accounts = (
        Account.query.filter_by(access_code_id=acc.id, type="Kas & Bank")
        .order_by(Account.code.asc())
        .all()
    )
    invoices = (
        SalesInvoice.query.filter_by(access_code_id=acc.id)
        .order_by(SalesInvoice.date.desc(), SalesInvoice.id.desc())
        .all()
    )

    if request.method == "POST":
        date_str = (request.form.get("date") or "").strip()
        invoice_id = (request.form.get("invoice_id") or "").strip()
        cash_code = (request.form.get("cash_account") or "").strip()
        amount_str = (request.form.get("amount") or "").strip()
        memo = (request.form.get("memo") or "").strip()

        if not date_str or not invoice_id or not cash_code or not amount_str:
            flash("Field wajib belum lengkap.", "error")
            return redirect(url_for("main.ar_payment_edit", pay_id=pay_id))

        inv = SalesInvoice.query.filter_by(access_code_id=acc.id, id=int(invoice_id)).first()
        if not inv:
            flash("Invoice tidak ditemukan.", "error")
            return redirect(url_for("main.ar_payment_edit", pay_id=pay_id))

        cash_acc = Account.query.filter_by(access_code_id=acc.id, code=cash_code).first()
        if not cash_acc:
            flash("Akun kas/bank tidak valid.", "error")
            return redirect(url_for("main.ar_payment_edit", pay_id=pay_id))

        try:
            amt = float(amount_str)
            if amt <= 0:
                raise ValueError()
        except ValueError:
            flash("Nominal harus angka > 0.", "error")
            return redirect(url_for("main.ar_payment_edit", pay_id=pay_id))

        pay.date = _parse_date(date_str)
        pay.invoice_id = inv.id
        pay.invoice_no = inv.invoice_no
        pay.cash_account_code = cash_acc.code
        pay.cash_account_name = cash_acc.name
        pay.amount = amt
        pay.memo = memo or None

        db.session.commit()
        _rebuild_everything()

        flash("Pembayaran piutang diupdate.", "success")
        return redirect(url_for("main.ar_payment_home"))

    return render_template("ar_payment_edit.html", pay=pay, cash_accounts=cash_accounts, invoices=invoices)


@bp.post("/ar/payments/<int:pay_id>/delete")
def ar_payment_delete(pay_id: int):
    acc = _require_access()
    if not acc:
        return redirect(url_for("main.enter_code"))

    pay = ARPayment.query.filter_by(access_code_id=acc.id, id=pay_id).first_or_404()
    db.session.delete(pay)
    db.session.commit()

    _rebuild_everything()

    flash("Pembayaran piutang dihapus.", "success")
    return redirect(url_for("main.ar_payment_home"))

